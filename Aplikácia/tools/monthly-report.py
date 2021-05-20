#!/usr/bin/env python3

import json
import sys
import openpyxl
from openpyxl.styles import PatternFill, Color
from openpyxl.comments import Comment
import re
from calendar import Calendar
from datetime import time, timedelta
from math import ceil
from os.path import dirname, join

def index_by(key_prop, array_of_dicts):
    indexed = {}
    for item in array_of_dicts:
        key = item[key_prop]
        if key in indexed:
            indexed[key].append(item)
        else:
            indexed[key] = [item]
    return indexed

def delta_from_time(t: time) -> timedelta:
    return timedelta(days=0,
        hours=t.hour, minutes=t.minute, seconds=t.second
    )

def delta_between_times(start: time, end: time) -> timedelta:
    return delta_from_time(end) - delta_from_time(start)

class Absence:
    SVIATOK = 'sviatok alebo iný deň voľna'
    DOVOLENKA = 'dovolenka'
    CHOROBA = 'choroba'
    URAZ = 'úraz alebo choroba z povolania'
    OSETROVANIE = 'ošetrovanie člena rodiny'
    OSOBNE_PREKAZKY = 'dôležité osobné prekážky v práci'
    CAKANIE = 'celodenné čakanie na prácu'
    VOLNO = 'pracovné voľno bez náhrady mzdy'
    NEOSPRAVEDLNENA = 'neospravedlnená absencia'
    MATERSKA = 'materská dovolenka'
    RODICOVSKA = 'rodičovská dovolenka'
    OSTATNE = 'ostatné'
    VSEOBECNE_PREKAZKY = 'prekážky z dôvodov všeobecného záujmu'
    NAHRADNE_VOLNO = 'náhradné voľno'
    SKOLENIE = 'školenie'
    STUDIUM = 'štúdium popri zamestnaní'
    PRACA_DOMA = 'práca doma'
    PRACOVNA_CESTA = 'pracovná cesta'
    INA_NEPRITOMNOST = 'iná neprítomnosť'

    counts_as_work = {
        SKOLENIE: True,
        STUDIUM: True,
        PRACA_DOMA: True,
        PRACOVNA_CESTA: True
    }

    from_web_type = {
        1: CHOROBA,
        2: PRACOVNA_CESTA,
        3: DOVOLENKA,
        4: PRACA_DOMA,
        5: INA_NEPRITOMNOST,
        6: MATERSKA,
        7: RODICOVSKA
    }

    abbreviation = {
        SVIATOK: 'S',
        DOVOLENKA: 'D',
        CHOROBA: 'CH',
        URAZ: 'Ú',
        OSETROVANIE: 'O',
        OSOBNE_PREKAZKY: 'P',
        CAKANIE: 'C',
        VOLNO: 'V',
        NEOSPRAVEDLNENA: 'A',
        MATERSKA: 'MD',
        RODICOVSKA: 'RD',
        OSTATNE: 'X',
        VSEOBECNE_PREKAZKY: 'Z',
        NAHRADNE_VOLNO: 'Nv',
        SKOLENIE: 'Šk',
        STUDIUM: 'Št'
    }

class OverviewSheetBuilder:
    TEMPLATE_FILENAME = 'monthly-report-template.xlsx'
    TEMPLATE_PATH = join(dirname(__file__), TEMPLATE_FILENAME)
    TEMPLATE_SHEET_TITLE = 'Šablóna'
    SHEET_TITLE_FORMAT = 'Hárok{}'
    MONTH_YEAR_CELL = 'A7'
    FIRST_ROW = 11
    ROWS_PER_SHEET = 30
    FIRST_DAY_COL = 4
    WORK_DAYS_COL = FIRST_DAY_COL + 31
    FIRST_ABSENCE_KIND_COL = WORK_DAYS_COL + 1
    ABSENCE_KINDS_ORDER = [
        Absence.SVIATOK,
        Absence.DOVOLENKA,
        Absence.CHOROBA,
        Absence.URAZ,
        Absence.OSETROVANIE,
        Absence.OSOBNE_PREKAZKY,
        Absence.CAKANIE,
        Absence.VOLNO,
        Absence.NEOSPRAVEDLNENA,
        Absence.MATERSKA,
        Absence.RODICOVSKA,
        Absence.OSTATNE
    ]
    TOTAL_ABSENCE_DAYS_COL = FIRST_ABSENCE_KIND_COL + len(ABSENCE_KINDS_ORDER)
    CATEGORIZE_MANUALLY_COL = TOTAL_ABSENCE_DAYS_COL + 1

    PUBLIC_HOLIDAY_ABSENCE_KIND = Absence.SVIATOK
    PUBLIC_HOLIDAY_FILL = PatternFill(
        patternType='solid',
        fgColor=Color(theme=0, tint=-0.35)
    )
    WEEKEND_ABSENCE_KIND = Absence.OSTATNE
    WEEKEND_FILL = PUBLIC_HOLIDAY_FILL

    PRESENCE_VALUE = '/'

    CATEGORIZE_MANUALLY_VALUE = '?'

    WORK_DAY_HOURS = 8
    WORK_DAY_FRACTION_DENOMINATOR = 4

    def __init__(self, data):
        self.year = data['year']
        self.month = data['month']
        self.month_sk = data['month_sk']
        self.personal_id_prefix = data['personal_id_prefix']
        self.employees = data['employees']
        self.public_holidays = index_by('day', data['public_holidays'])
        self.absences = dict((eid, index_by('day', eabs))
            for (eid, eabs) in index_by('user_id', data['absences']).items()
        )

    def build(self, output):
        wb = openpyxl.load_workbook(self.TEMPLATE_PATH)
        OverviewSheetBuilder.iterate(
            lambda: self.fill_in_book(iter(self.employees), wb)
        )
        del wb[self.TEMPLATE_SHEET_TITLE]
        wb.save(output)

    @staticmethod
    def iterate(procedure):
        try:
            procedure()
            return True
        except StopIteration:
            return False

    def fill_in_book(self, employees, wb):
        sheet_num = 0
        cont = True
        while cont:
            sheet_num += 1
            ws = wb.copy_worksheet(wb[self.TEMPLATE_SHEET_TITLE])
            ws.conditional_formatting = \
                wb[self.TEMPLATE_SHEET_TITLE].conditional_formatting
            ws.title = self.SHEET_TITLE_FORMAT.format(sheet_num)
            ws[self.MONTH_YEAR_CELL] = \
                ws[self.MONTH_YEAR_CELL].value.format(
                    month = self.month,
                    month_sk = self.month_sk,
                    year = self.year
                )
            cont = OverviewSheetBuilder.iterate(
                lambda: self.fill_in_sheet(employees, ws)
            )

    def fill_in_sheet(self, employees, ws):
        prefix_re = re.compile('^' + self.personal_id_prefix)
        nat_re = re.compile('^[0-9]+$')
        for row in range(self.FIRST_ROW, self.FIRST_ROW + self.ROWS_PER_SHEET):
            employee = next(employees)
            short_id = prefix_re.sub('', str(employee['personal_id']))
            full_name = '{} {}'.format(employee['surname'], employee['name'])
            ws.cell(row, 1).value = int(short_id) if nat_re.match(short_id) else short_id
            ws.cell(row, 2).value = full_name
            (presence_days, absence_totals, categorize_manually) = \
                self.fill_in_days(row, employee['id'], full_name, ws)
            self.fill_in_totals(row, presence_days, absence_totals,
                                categorize_manually, ws)

    def fill_in_days(self, row, employee_id, full_name, ws):
        presence_days = 0
        absence_totals = self.init_absence_totals()
        categorize_manually = 0
        absences = (self.absences[employee_id]
            if employee_id in self.absences
            else {})
        for dayOfMonth, dayOfWeek \
                in Calendar().itermonthdays2(self.year, self.month):
            if dayOfMonth < 1:
                continue
            cell = ws.cell(row, self.FIRST_DAY_COL + dayOfMonth - 1)
            if dayOfWeek >= 5:
                cell.value = Absence.abbreviation[self.WEEKEND_ABSENCE_KIND]
                cell.fill = self.WEEKEND_FILL
                continue
            if dayOfMonth in self.public_holidays:
                cell.value = Absence.abbreviation[self.PUBLIC_HOLIDAY_ABSENCE_KIND]
                cell.fill = self.PUBLIC_HOLIDAY_FILL
                absence_totals[self.PUBLIC_HOLIDAY_ABSENCE_KIND] += 1
                continue
            if dayOfMonth in absences:
                absence = absences[dayOfMonth][0]
                kind = Absence.from_web_type[absence['type']]
                duration = self.work_day_fraction_from_delta(
                    delta_between_times(
                        time.fromisoformat(absence['from_time']),
                        time.fromisoformat(absence['to_time'])
                    )
                )
                maybe_presence = (self.PRESENCE_VALUE if duration < 1
                                    else '');
                if kind in Absence.abbreviation:
                    cell.value = maybe_presence + Absence.abbreviation[kind]
                elif kind in Absence.counts_as_work:
                    cell.value = self.PRESENCE_VALUE
                else:
                    cell.value = maybe_presence + self.CATEGORIZE_MANUALLY_VALUE
                    cell.comment = Comment(
                        '{}: {}'.format(kind, absence['description']),
                        full_name
                    )
                if kind in Absence.counts_as_work:
                    presence_days += duration
                elif kind in absence_totals:
                    absence_totals[kind] += duration
                else:
                    categorize_manually += duration
                presence_days += 1 - duration
                continue
            cell.value = self.PRESENCE_VALUE
            presence_days += 1
        return (presence_days, absence_totals, categorize_manually)

    def fill_in_totals(self, row, presence_days, absence_totals,
                       categorize_manually, ws):
        ws.cell(row, self.WORK_DAYS_COL).value = presence_days
        col = self.FIRST_ABSENCE_KIND_COL
        for kind in self.ABSENCE_KINDS_ORDER:
            if absence_totals[kind] != 0:
                ws.cell(row, col).value = absence_totals[kind]
            col += 1
        ws.cell(row, self.TOTAL_ABSENCE_DAYS_COL).value = \
            '=SUM({}:{})'.format(
                ws.cell(row, self.FIRST_ABSENCE_KIND_COL).coordinate,
                ws.cell(row, self.FIRST_ABSENCE_KIND_COL +
                        len(self.ABSENCE_KINDS_ORDER) - 1).coordinate)
        if categorize_manually > 0:
            ws.cell(row, self.CATEGORIZE_MANUALLY_COL).value =\
                '{}?'.format(categorize_manually)

    def init_absence_totals(self):
        return dict((absenceType, 0)
            for absenceType in self.ABSENCE_KINDS_ORDER)

    def work_day_fraction_from_delta(self, d: timedelta) -> float:
        return (ceil(d.total_seconds() * self.WORK_DAY_FRACTION_DENOMINATOR
                        / (self.WORK_DAY_HOURS * 3600))
                / self.WORK_DAY_FRACTION_DENOMINATOR)

def main():
    if len(sys.argv) != 3:
        sys.stderr.write(
            "Usage: {} report-data.json output-workbook.xlsx\n"
            .format(sys.argv[0])
        )
        exit(1)
    with open(sys.argv[1], "r") as input:
        OverviewSheetBuilder(json.load(input)).build(sys.argv[2])

if __name__ == "__main__":
    main()
